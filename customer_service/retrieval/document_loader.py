from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterable
from zipfile import BadZipFile

from customer_service.retrieval.schemas import KnowledgeRecord

try:
    from pypdf import PdfReader
except Exception:  # pragma: no cover
    PdfReader = None

try:
    from docx import Document as WordDocument
except Exception:  # pragma: no cover
    WordDocument = None

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None


AGENT_BY_DOMAIN = {
    "technical": "technical_expert",
    "sales": "sales_expert",
    "support": "support_expert",
    "feedback": "feedback_expert",
    "share": "shared",
}

SHARED_DOMAIN = "share"
SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".xlsx", ".csv", ".txt", ".md"}


def chunk_text(text: str, chunk_size: int = 600, overlap: int = 120) -> list[str]:
    normalized = "\n".join([line.strip() for line in text.splitlines() if line.strip()])
    if not normalized:
        return []
    if len(normalized) <= chunk_size:
        return [normalized]

    chunks: list[str] = []
    start = 0
    step = max(1, chunk_size - overlap)
    while start < len(normalized):
        chunk = normalized[start : start + chunk_size].strip()
        if chunk:
            chunks.append(chunk)
        start += step
    return chunks


def _read_text_file(path: Path) -> str:
    for encoding in ("utf-8", "utf-8-sig", "gb18030"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    return path.read_text(encoding="utf-8", errors="ignore")


def _load_pdf(path: Path) -> Iterable[tuple[str, dict[str, object]]]:
    if PdfReader is None:
        raise RuntimeError("未安装 pypdf，无法解析 PDF 文件。")
    reader = PdfReader(str(path))
    for index, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        yield text, {"page": index}


def _load_docx(path: Path) -> Iterable[tuple[str, dict[str, object]]]:
    if WordDocument is None:
        raise RuntimeError("未安装 python-docx，无法解析 Word 文件。")
    try:
        document = WordDocument(str(path))
    except BadZipFile as exc:
        raise RuntimeError(f"Word 文件损坏或格式异常：{path.name}") from exc
    paragraphs = [paragraph.text.strip() for paragraph in document.paragraphs if paragraph.text.strip()]
    yield "\n".join(paragraphs), {}


def _load_xlsx(path: Path) -> Iterable[tuple[str, dict[str, object]]]:
    if load_workbook is None:
        raise RuntimeError("未安装 openpyxl，无法解析 Excel 文件。")
    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    for sheet in workbook.worksheets:
        rows: list[str] = []
        for row in sheet.iter_rows(values_only=True):
            values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
            if values:
                rows.append(" | ".join(values))
        yield "\n".join(rows), {"sheet": sheet.title}


def _load_csv(path: Path) -> Iterable[tuple[str, dict[str, object]]]:
    rows: list[str] = []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        for row in reader:
            values = [cell.strip() for cell in row if cell and cell.strip()]
            if values:
                rows.append(" | ".join(values))
    yield "\n".join(rows), {}


def _load_text_like(path: Path) -> Iterable[tuple[str, dict[str, object]]]:
    yield _read_text_file(path), {}


def load_raw_documents(raw_dir: str | Path) -> list[KnowledgeRecord]:
    base_dir = Path(raw_dir)
    if not base_dir.exists():
        return []

    records: list[KnowledgeRecord] = []
    for file_path in sorted(base_dir.rglob("*")):
        if not file_path.is_file() or file_path.suffix.lower() not in SUPPORTED_EXTENSIONS:
            continue

        domain = file_path.parent.name.lower()
        if domain not in AGENT_BY_DOMAIN:
            continue

        loader = {
            ".pdf": _load_pdf,
            ".docx": _load_docx,
            ".xlsx": _load_xlsx,
            ".csv": _load_csv,
            ".txt": _load_text_like,
            ".md": _load_text_like,
        }[file_path.suffix.lower()]

        for section_index, (text, section_meta) in enumerate(loader(file_path)):
            for chunk_index, chunk in enumerate(chunk_text(text), start=1):
                source_stem = file_path.stem.replace(" ", "_").lower()
                section_label = section_meta.get("page") or section_meta.get("sheet") or section_index + 1
                record_id = f"{domain}_{source_stem}_{section_label}_{chunk_index}"
                metadata = {
                    "domain": domain,
                    "agent": AGENT_BY_DOMAIN[domain],
                    "source": file_path.name,
                    "source_path": str(file_path).replace('\\', '/'),
                    "priority": 7,
                    "keywords": [],
                    "chunk_index": chunk_index,
                    "is_shared": domain == SHARED_DOMAIN,
                    **section_meta,
                }
                records.append(KnowledgeRecord(id=record_id, content=chunk, metadata=metadata))
    return records